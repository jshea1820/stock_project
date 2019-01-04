import openpyxl as xl
import pandas as pd
import numpy as np

def excel_index(integer_value):
    '''generates a letter index based on a given integer value'''
    ''' 0->A, 10->K, 26->AA, etc'''
    '''works up with values <= 26^2 - 1'''

    if integer_value < 26:
        return chr(65 + integer_value)

    else:
        first_num = 64 + int(integer_value / 26)
        second_num = 65 + (integer_value % 26)

    return chr(first_num) + chr(second_num)



# Opens opens stock data
data = pd.read_csv('data_stocks.csv')
data = data.values

# Opens excel workbook to write in
wb = xl.load_workbook("dailyDeltas.xlsx")
ws = wb.get_sheet_by_name("Sheet1")

# Loops through data_stocks gathering change data
last_time = 0
day = 1
for time_index, time in enumerate(data[:,0]):
    #print("Time: " + str(time_index) + ", " + str(time))

    if last_time == 0: # handles case of first row
        #print("Parsing first row")
        open_prices = []
        for i in range(2,502):
            open_prices.append(data[time_index, i])
            
        
    elif time - last_time != 60: # new day detected
        #print("New day detected: " + str(time_index) + " " + str(time))

        # Records all the last day closing prices and calculates change
        for i in range(2,502):
            #print("Open price is " + str(open_prices[i-2]))
            #print("Close price is " + str(data[time_index - 1, i]))
            close_price = data[time_index - 1, i]
            #print("Delta is " + str(close_price - open_prices[i-2]))
            delta = close_price - open_prices[i-2]
            #print("Percentage change is " + str(delta / open_prices[i-2]))
            delta_p = delta / open_prices[i-2]
            ws[excel_index(i-1) + str(day + 1)].value = delta_p
            open_prices[i-2] = data[time_index, i]

        if day != 1:
            market_change = data[time_index - 1, 1] - market_open
            if market_change > 0:
                ws['SH' + str(day)].value = 1
            else:
                ws['SH' + str(day)].value = 0
        
        market_open = data[time_index, 1]
       
        day += 1

            
            
    last_time = time

wb.save("dailyDeltas.xlsx")




